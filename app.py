import csv
import io
import os
from calendar import month_abbr
from datetime import date, datetime

from dotenv import load_dotenv
from flask import Flask, Response, flash, redirect, render_template, request, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from sqlalchemy import func, inspect, text
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from models import Expense, Income, Item, Profile, Purchase, PurchaseLine, Sale, StockLog, Supplier, db

ALLOWED_AVATAR_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

MOVEMENT_TYPE_LABELS = {
    "sale": "Sale",
    "purchase": "Purchase",
    "restock": "Restock",
    "sponsor": "Sponsor/Influencer",
    "damaged": "Damaged",
    "personal_use": "Personal Use",
    "sample_gift": "Sample/Gift",
    "initial_stock": "Initial Stock",
    "other": "Other",
}

load_dotenv()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")

database_url = os.environ.get("DATABASE_URL", "").strip()
if database_url:
    # Render/Heroku style URLs sometimes start with postgres:// which SQLAlchemy no longer accepts.
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
else:
    os.makedirs(os.path.join(app.root_path, "data"), exist_ok=True)
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(
        app.root_path, "data", "invenex.db"
    )
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024  # 2MB upload limit

UPLOAD_FOLDER = os.path.join(app.root_path, "static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.init_app(app)

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_DISPLAY_NAME = os.environ.get("ADMIN_DISPLAY_NAME", ADMIN_USERNAME)
ADMIN_PASSWORD_HASH = generate_password_hash(
    os.environ.get("ADMIN_PASSWORD", "admin")
)


class AdminUser(UserMixin):
    id = "admin"


@login_manager.user_loader
def load_user(user_id):
    if user_id == "admin":
        return AdminUser()
    return None


def get_profile():
    profile = db.session.get(Profile, 1)
    if profile is None:
        profile = Profile(id=1, display_name=ADMIN_DISPLAY_NAME, avatar_filename=None)
        db.session.add(profile)
        db.session.commit()
    return profile


@app.context_processor
def inject_globals():
    profile = get_profile()
    return {
        "admin_display_name": profile.display_name or ADMIN_DISPLAY_NAME,
        "admin_avatar": profile.avatar_filename,
        "current_year": date.today().year,
        "movement_type_labels": MOVEMENT_TYPE_LABELS,
    }


def _add_missing_columns():
    """Lightweight in-place migration: add newly introduced columns to existing tables."""
    inspector = inspect(db.engine)
    required = {
        "items": [
            ("category", "VARCHAR(80)"),
            ("sku", "VARCHAR(60)"),
            ("buying_price", "FLOAT"),
            ("selling_price", "FLOAT"),
            ("discount_price", "FLOAT"),
        ],
        "income": [("category", "VARCHAR(80)")],
        "expenses": [("category", "VARCHAR(80)")],
        "stock_logs": [
            ("movement_type", "VARCHAR(30)"),
            ("supplier", "VARCHAR(120)"),
            ("po_number", "VARCHAR(60)"),
            ("po_batch", "VARCHAR(60)"),
        ],
        "sales": [("stock_log_id", "INTEGER")],
    }
    for table, columns in required.items():
        if table not in inspector.get_table_names():
            continue
        existing = {col["name"] for col in inspector.get_columns(table)}
        for column_name, column_type in columns:
            if column_name not in existing:
                db.session.execute(
                    text(f"ALTER TABLE {table} ADD COLUMN {column_name} {column_type}")
                )
    db.session.commit()


with app.app_context():
    db.create_all()
    _add_missing_columns()


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and check_password_hash(
            ADMIN_PASSWORD_HASH, password
        ):
            login_user(AdminUser())
            return redirect(url_for("dashboard"))
        flash("Username ba password vul!", "danger")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/profile")
@login_required
def profile():
    return render_template("profile.html", profile=get_profile())


@app.route("/profile/update", methods=["POST"])
@login_required
def update_profile():
    profile_row = get_profile()

    display_name = request.form.get("display_name", "").strip()
    if display_name:
        profile_row.display_name = display_name

    avatar = request.files.get("avatar")
    if avatar and avatar.filename:
        ext = avatar.filename.rsplit(".", 1)[-1].lower() if "." in avatar.filename else ""
        if ext not in ALLOWED_AVATAR_EXTENSIONS:
            flash("Shudhu image file (png, jpg, jpeg, gif, webp) upload kora jabe.", "danger")
            return redirect(url_for("profile"))

        filename = secure_filename(f"avatar_{int(datetime.utcnow().timestamp())}.{ext}")
        old_filename = profile_row.avatar_filename
        avatar.save(os.path.join(UPLOAD_FOLDER, filename))
        profile_row.avatar_filename = filename

        if old_filename:
            old_path = os.path.join(UPLOAD_FOLDER, old_filename)
            if os.path.exists(old_path):
                os.remove(old_path)

    db.session.commit()
    flash("Profile update kora hoyeche.", "success")
    return redirect(url_for("profile"))


@app.errorhandler(413)
def file_too_large(e):
    flash("File size 2MB er beshi hote parbe na.", "danger")
    return redirect(url_for("profile"))


def _month_sum(model, year, month):
    start = _month_start(year, month)
    next_y, next_m = _next_month(year, month)
    end_exclusive = _month_start(next_y, next_m)
    return db.session.query(func.coalesce(func.sum(model.amount), 0)).filter(
        model.entry_date >= start, model.entry_date < end_exclusive
    ).scalar()


def _trend(current, previous):
    if previous == 0:
        if current == 0:
            return {"direction": "flat", "pct": None}
        return {"direction": "up", "pct": None}
    pct = (current - previous) / previous * 100
    if abs(pct) < 1:
        return {"direction": "flat", "pct": pct}
    return {"direction": "up" if pct > 0 else "down", "pct": pct}


@app.route("/")
@login_required
def dashboard():
    total_income = db.session.query(func.coalesce(func.sum(Income.amount), 0)).scalar()
    total_expense = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).scalar()
    balance = total_income - total_expense

    today = date.today()
    prev_year, prev_month = (today.year - 1, 12) if today.month == 1 else (today.year, today.month - 1)
    this_month_income = _month_sum(Income, today.year, today.month)
    last_month_income = _month_sum(Income, prev_year, prev_month)
    this_month_expense = _month_sum(Expense, today.year, today.month)
    last_month_expense = _month_sum(Expense, prev_year, prev_month)
    income_trend = _trend(this_month_income, last_month_income)
    expense_trend = _trend(this_month_expense, last_month_expense)

    total_items = Item.query.count()
    low_stock_items = [item for item in Item.query.all() if item.is_low_stock]

    recent_income = Income.query.order_by(Income.entry_date.desc(), Income.id.desc()).limit(5).all()
    recent_expense = Expense.query.order_by(Expense.entry_date.desc(), Expense.id.desc()).limit(5).all()
    recent_stock_logs = StockLog.query.order_by(StockLog.id.desc()).limit(5).all()
    recent_sales = Sale.query.order_by(Sale.sale_date.desc(), Sale.id.desc()).limit(5).all()

    due_sales = Sale.query.filter_by(status="due").order_by(Sale.sale_date.desc(), Sale.id.desc()).all()
    due_total = sum(sale.amount for sale in due_sales)

    total_stock_remaining = db.session.query(func.coalesce(func.sum(Item.quantity), 0)).scalar()
    total_stock_out_all_time = db.session.query(func.coalesce(func.sum(StockLog.quantity), 0)).filter(
        StockLog.change_type == "out"
    ).scalar()
    # "Initial" = current remaining + everything that has ever left (sold/given away/damaged).
    # This stays correct even for items created before stock movements started being logged.
    total_stock_initial = total_stock_remaining + total_stock_out_all_time
    total_units_sold = db.session.query(func.coalesce(func.sum(Sale.quantity), 0)).scalar()
    total_sponsor_given = db.session.query(func.coalesce(func.sum(StockLog.quantity), 0)).filter(
        StockLog.change_type == "out", StockLog.movement_type == "sponsor"
    ).scalar()

    return render_template(
        "dashboard.html",
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        income_trend=income_trend,
        expense_trend=expense_trend,
        total_items=total_items,
        low_stock_items=low_stock_items,
        recent_income=recent_income,
        recent_expense=recent_expense,
        recent_stock_logs=recent_stock_logs,
        recent_sales=recent_sales,
        due_sales=due_sales,
        due_total=due_total,
        total_stock_remaining=total_stock_remaining,
        total_stock_initial=total_stock_initial,
        total_units_sold=total_units_sold,
        total_sponsor_given=total_sponsor_given,
    )


# ---------- Inventory ----------

def _filtered_items():
    query = Item.query
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    low_stock_only = request.args.get("low_stock") == "1"

    if q:
        query = query.filter(Item.name.ilike(f"%{q}%"))
    if category:
        query = query.filter(Item.category == category)

    items = query.order_by(Item.name).all()
    if low_stock_only:
        items = [item for item in items if item.is_low_stock]
    return items


def _filtered_stock_logs():
    query = StockLog.query.join(Item)
    q = request.args.get("stock_q", "").strip()
    change_type = request.args.get("stock_type", "").strip()
    movement_type = request.args.get("stock_category", "").strip()
    date_from = request.args.get("stock_date_from", "").strip()
    date_to = request.args.get("stock_date_to", "").strip()

    if q:
        query = query.filter(db.or_(StockLog.reason.ilike(f"%{q}%"), Item.name.ilike(f"%{q}%")))
    if change_type in ("in", "out"):
        query = query.filter(StockLog.change_type == change_type)
    if movement_type in MOVEMENT_TYPE_LABELS:
        query = query.filter(StockLog.movement_type == movement_type)
    if date_from:
        query = query.filter(StockLog.entry_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(StockLog.entry_date <= date.fromisoformat(date_to))

    return query.order_by(StockLog.entry_date.desc(), StockLog.id.desc()).all()


@app.route("/inventory")
@login_required
def inventory():
    items = _filtered_items()
    all_items = Item.query.order_by(Item.name).all()
    categories = [
        row[0] for row in
        db.session.query(Item.category).filter(Item.category.isnot(None)).distinct().order_by(Item.category)
    ]
    return render_template(
        "inventory.html",
        items=items,
        all_items=all_items,
        categories=categories,
        today=date.today().isoformat(),
        filters=request.args,
    )


@app.route("/inventory/sales")
@login_required
def sale_page():
    sales = Sale.query.order_by(Sale.sale_date.desc(), Sale.id.desc()).all()
    all_items = Item.query.order_by(Item.name).all()
    return render_template(
        "sale_page.html",
        sales=sales,
        all_items=all_items,
        today=date.today().isoformat(),
    )


@app.route("/inventory/stock-report")
@login_required
def stock_report_page():
    stock_logs = _filtered_stock_logs()
    return render_template(
        "stock_report.html",
        stock_logs=stock_logs,
        filters=request.args,
    )


@app.route("/inventory/stock-history/export")
@login_required
def export_stock_history():
    logs = _filtered_stock_logs()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Item", "Type", "Category", "Quantity", "Reason", "Supplier", "PO Number", "PO Batch"])
    for log in logs:
        writer.writerow([
            log.entry_date,
            log.item.name if log.item else "",
            log.change_type.upper(),
            MOVEMENT_TYPE_LABELS.get(log.movement_type, ""),
            log.quantity,
            log.reason or "",
            log.supplier or "",
            log.po_number or "",
            log.po_batch or "",
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=stock_history.csv"},
    )


@app.route("/inventory/export")
@login_required
def export_inventory():
    items = _filtered_items()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Item", "SKU", "Category", "Unit", "Current Quantity", "Low Stock Threshold", "Buying Price", "Regular Sale Price", "Discount Sale Price"])
    for item in items:
        writer.writerow([
            item.name,
            item.sku or "",
            item.category or "",
            item.unit,
            item.quantity,
            item.low_stock_threshold,
            item.buying_price if item.buying_price is not None else "",
            item.selling_price if item.selling_price is not None else "",
            item.discount_price if item.discount_price is not None else "",
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=inventory.csv"},
    )


@app.route("/inventory/add", methods=["POST"])
@login_required
def add_item():
    name = request.form.get("name", "").strip()
    sku = request.form.get("sku", "").strip()
    category = request.form.get("category", "").strip()
    unit = request.form.get("unit", "pcs").strip() or "pcs"
    quantity = request.form.get("quantity", "0")
    threshold = request.form.get("low_stock_threshold", "5")
    buying_price = request.form.get("buying_price", "").strip()
    selling_price = request.form.get("selling_price", "").strip()
    discount_price = request.form.get("discount_price", "").strip()
    supplier = request.form.get("supplier", "").strip()
    po_number = request.form.get("po_number", "").strip()
    po_batch = request.form.get("po_batch", "").strip()

    if not name:
        flash("Item er naam dite hobe.", "danger")
        return redirect(url_for("inventory"))

    if sku and Item.query.filter(func.lower(Item.sku) == sku.lower()).first():
        flash(f"SKU '{sku}' already ekta item-e use hoyeche. Onno SKU diye try korun.", "danger")
        return redirect(url_for("inventory"))

    quantity_int = int(quantity or 0)
    item = Item(
        name=name,
        sku=sku or None,
        category=category or None,
        unit=unit,
        quantity=quantity_int,
        low_stock_threshold=int(threshold or 5),
        buying_price=float(buying_price) if buying_price else None,
        selling_price=float(selling_price) if selling_price else None,
        discount_price=float(discount_price) if discount_price else None,
    )
    db.session.add(item)
    db.session.flush()

    if quantity_int > 0:
        db.session.add(StockLog(
            item_id=item.id,
            change_type="in",
            movement_type="initial_stock",
            quantity=quantity_int,
            reason="Initial stock",
            supplier=supplier or None,
            po_number=po_number or None,
            po_batch=po_batch or None,
            entry_date=date.today(),
        ))

    db.session.commit()
    flash(f"'{name}' item add kora hoyeche.", "success")
    return redirect(url_for("inventory"))


@app.route("/inventory/<int:item_id>/stock", methods=["POST"])
@login_required
def update_stock(item_id):
    item = Item.query.get_or_404(item_id)
    change_type = request.form.get("change_type")
    quantity = int(request.form.get("quantity", "0") or 0)
    reason = request.form.get("reason", "").strip()
    entry_date = request.form.get("entry_date") or date.today().isoformat()

    if quantity <= 0:
        flash("Quantity 0 er cheye beshi hote hobe.", "danger")
        return redirect(url_for("inventory"))

    if change_type == "in":
        item.quantity += quantity
    elif change_type == "out":
        item.quantity -= quantity
    else:
        flash("Invalid stock type.", "danger")
        return redirect(url_for("inventory"))

    log = StockLog(
        item_id=item.id,
        change_type=change_type,
        movement_type="restock" if change_type == "in" else "other",
        quantity=quantity,
        reason=reason,
        entry_date=date.fromisoformat(entry_date),
    )
    db.session.add(log)
    db.session.commit()
    flash(f"'{item.name}' er stock update hoyeche.", "success")
    return redirect(url_for("inventory"))


@app.route("/inventory/stock-out", methods=["POST"])
@login_required
def quick_stock_out():
    item = Item.query.get_or_404(int(request.form.get("item_id")))
    quantity = int(request.form.get("quantity", "0") or 0)
    movement_type = request.form.get("movement_type", "other")
    if movement_type not in MOVEMENT_TYPE_LABELS:
        movement_type = "other"
    note = request.form.get("note", "").strip()
    entry_date = request.form.get("entry_date") or date.today().isoformat()

    if quantity <= 0:
        flash("Quantity 0 er cheye beshi hote hobe.", "danger")
        return redirect(url_for("inventory"))

    reason = f"{MOVEMENT_TYPE_LABELS[movement_type]} - {note}" if note else MOVEMENT_TYPE_LABELS[movement_type]

    item.quantity -= quantity
    db.session.add(StockLog(
        item_id=item.id,
        change_type="out",
        movement_type=movement_type,
        quantity=quantity,
        reason=reason,
        entry_date=date.fromisoformat(entry_date),
    ))
    db.session.commit()
    flash(f"'{item.name}' theke stock out kora hoyeche.", "success")
    return redirect(url_for("inventory"))


@app.route("/inventory/<int:item_id>/delete", methods=["POST"])
@login_required
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash(f"'{item.name}' delete kora hoyeche.", "info")
    return redirect(url_for("inventory"))


# ---------- Sales ----------

def _create_income_for_sale(sale, item):
    income = Income(
        purpose=f"Sale - {item.name if item else 'Item'}",
        category="Sale",
        amount=sale.amount,
        entry_date=sale.sale_date,
        note=sale.customer_name,
    )
    db.session.add(income)
    db.session.flush()
    sale.income_id = income.id


@app.route("/sales/add", methods=["POST"])
@login_required
def add_sale():
    item = Item.query.get_or_404(int(request.form.get("item_id")))
    quantity = int(request.form.get("quantity", "0") or 0)
    amount = float(request.form.get("amount", "0") or 0)
    customer_name = request.form.get("customer_name", "").strip()
    status = request.form.get("status", "due")
    if status not in ("due", "complete"):
        status = "due"
    sale_date = request.form.get("sale_date") or date.today().isoformat()
    note = request.form.get("note", "").strip()

    if quantity <= 0 or amount <= 0:
        flash("Sale quantity ar amount 0 er cheye beshi hote hobe.", "danger")
        return redirect(url_for("sale_page"))

    sale_date_obj = date.fromisoformat(sale_date)
    item.quantity -= quantity

    stock_log = StockLog(
        item_id=item.id,
        change_type="out",
        movement_type="sale",
        quantity=quantity,
        reason=f"Sale - {customer_name or 'Customer'}",
        entry_date=sale_date_obj,
    )
    db.session.add(stock_log)
    db.session.flush()

    sale = Sale(
        item_id=item.id,
        customer_name=customer_name,
        quantity=quantity,
        amount=amount,
        status=status,
        sale_date=sale_date_obj,
        note=note,
        stock_log_id=stock_log.id,
    )
    db.session.add(sale)
    db.session.flush()

    if status == "complete":
        _create_income_for_sale(sale, item)

    db.session.commit()
    flash("Sale record kora hoyeche.", "success")
    return redirect(url_for("sale_page"))


@app.route("/sales/<int:sale_id>/update", methods=["POST"])
@login_required
def update_sale(sale_id):
    sale = Sale.query.get_or_404(sale_id)

    sale.customer_name = request.form.get("customer_name", "").strip()
    amount = request.form.get("amount")
    if amount:
        sale.amount = float(amount)
    sale.note = request.form.get("note", "").strip()
    sale_date = request.form.get("sale_date")
    if sale_date:
        sale.sale_date = date.fromisoformat(sale_date)

    quantity = request.form.get("quantity")
    if quantity:
        new_quantity = int(quantity)
        if new_quantity > 0 and new_quantity != sale.quantity:
            item = Item.query.get(sale.item_id)
            if item:
                item.quantity += sale.quantity - new_quantity
            if sale.stock_log_id:
                stock_log = StockLog.query.get(sale.stock_log_id)
                if stock_log:
                    stock_log.quantity = new_quantity
                    stock_log.entry_date = sale.sale_date
            sale.quantity = new_quantity

    new_status = request.form.get("status", sale.status)
    if new_status == "complete" and sale.status == "due":
        item = Item.query.get(sale.item_id)
        _create_income_for_sale(sale, item)
        sale.status = "complete"
    elif new_status == "due" and sale.status == "complete":
        if sale.income_id:
            income = Income.query.get(sale.income_id)
            if income:
                db.session.delete(income)
            sale.income_id = None
        sale.status = "due"
    elif sale.status == "complete" and sale.income_id:
        # Keep the linked income entry in sync with any edited amount/date/customer.
        income = Income.query.get(sale.income_id)
        if income:
            income.amount = sale.amount
            income.entry_date = sale.sale_date
            income.note = sale.customer_name

    db.session.commit()
    flash("Sale update kora hoyeche.", "success")
    return redirect(url_for("sale_page"))


@app.route("/sales/<int:sale_id>/delete", methods=["POST"])
@login_required
def delete_sale(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    item = Item.query.get(sale.item_id)
    if item:
        item.quantity += sale.quantity
    if sale.income_id:
        income = Income.query.get(sale.income_id)
        if income:
            db.session.delete(income)
    if sale.stock_log_id:
        stock_log = StockLog.query.get(sale.stock_log_id)
        if stock_log:
            db.session.delete(stock_log)
    db.session.delete(sale)
    db.session.commit()
    flash("Sale entry delete kora hoyeche.", "info")
    return redirect(url_for("sale_page"))


# ---------- Purchases ----------

def _create_expense_for_purchase(purchase):
    expense = Expense(
        purpose=f"Purchase - {purchase.supplier.name if purchase.supplier else 'Supplier'}",
        category="Purchase",
        amount=purchase.total_amount,
        entry_date=purchase.purchase_date,
        note=f"PO: {purchase.po_number}" if purchase.po_number else None,
    )
    db.session.add(expense)
    db.session.flush()
    purchase.expense_id = expense.id


@app.route("/suppliers/add", methods=["POST"])
@login_required
def add_supplier():
    name = request.form.get("name", "").strip()
    contact = request.form.get("contact", "").strip()
    notes = request.form.get("notes", "").strip()

    if not name:
        flash("Supplier naam dite hobe.", "danger")
        return redirect(url_for("purchase_page"))

    if Supplier.query.filter(func.lower(Supplier.name) == name.lower()).first():
        flash(f"Supplier '{name}' already ache.", "danger")
        return redirect(url_for("purchase_page"))

    db.session.add(Supplier(name=name, contact=contact or None, notes=notes or None))
    db.session.commit()
    flash(f"Supplier '{name}' add kora hoyeche.", "success")
    return redirect(url_for("purchase_page"))


@app.route("/purchases")
@login_required
def purchase_page():
    purchases = Purchase.query.order_by(Purchase.purchase_date.desc(), Purchase.id.desc()).all()
    suppliers = Supplier.query.order_by(Supplier.name).all()
    all_items = Item.query.order_by(Item.name).all()
    due_total = sum(p.total_amount for p in purchases if p.status == "due")
    return render_template(
        "purchase_page.html",
        purchases=purchases,
        suppliers=suppliers,
        all_items=all_items,
        due_total=due_total,
        today=date.today().isoformat(),
    )


@app.route("/purchases/add", methods=["POST"])
@login_required
def add_purchase():
    supplier_id = request.form.get("supplier_id")
    if not supplier_id:
        flash("Supplier select korte hobe.", "danger")
        return redirect(url_for("purchase_page"))
    supplier = Supplier.query.get_or_404(int(supplier_id))

    po_number = request.form.get("po_number", "").strip()
    po_batch = request.form.get("po_batch", "").strip()
    purchase_date = request.form.get("purchase_date") or date.today().isoformat()
    status = request.form.get("status", "due")
    if status not in ("due", "paid"):
        status = "due"
    note = request.form.get("note", "").strip()
    purchase_date_obj = date.fromisoformat(purchase_date)

    item_ids = request.form.getlist("item_id[]")
    quantities = request.form.getlist("quantity[]")
    unit_prices = request.form.getlist("unit_price[]")

    lines_data = []
    for item_id, qty, price in zip(item_ids, quantities, unit_prices):
        if not item_id or not qty or not price:
            continue
        qty_int = int(qty)
        price_float = float(price)
        if qty_int <= 0 or price_float < 0:
            continue
        lines_data.append((int(item_id), qty_int, price_float))

    if not lines_data:
        flash("Kom pokkhe ekta valid item line dite hobe (item, quantity, price).", "danger")
        return redirect(url_for("purchase_page"))

    purchase = Purchase(
        supplier_id=supplier.id,
        po_number=po_number or None,
        po_batch=po_batch or None,
        purchase_date=purchase_date_obj,
        status=status,
        note=note or None,
        total_amount=0,
    )
    db.session.add(purchase)
    db.session.flush()

    total_amount = 0.0
    for item_id, qty_int, price_float in lines_data:
        item = Item.query.get(item_id)
        if not item:
            continue
        item.quantity += qty_int
        subtotal = qty_int * price_float
        total_amount += subtotal

        stock_log = StockLog(
            item_id=item.id,
            change_type="in",
            movement_type="purchase",
            quantity=qty_int,
            reason=f"Purchase - {supplier.name}",
            supplier=supplier.name,
            po_number=po_number or None,
            po_batch=po_batch or None,
            entry_date=purchase_date_obj,
        )
        db.session.add(stock_log)
        db.session.flush()

        db.session.add(PurchaseLine(
            purchase_id=purchase.id,
            item_id=item.id,
            quantity=qty_int,
            unit_price=price_float,
            stock_log_id=stock_log.id,
        ))

    purchase.total_amount = round(total_amount, 2)

    if status == "paid":
        _create_expense_for_purchase(purchase)

    db.session.commit()
    flash("Purchase record kora hoyeche.", "success")
    return redirect(url_for("purchase_page"))


@app.route("/purchases/<int:purchase_id>/update", methods=["POST"])
@login_required
def update_purchase(purchase_id):
    purchase = Purchase.query.get_or_404(purchase_id)

    supplier_id = request.form.get("supplier_id")
    if supplier_id:
        purchase.supplier_id = int(supplier_id)
    purchase.po_number = request.form.get("po_number", "").strip() or None
    purchase.po_batch = request.form.get("po_batch", "").strip() or None
    purchase.note = request.form.get("note", "").strip() or None
    purchase_date = request.form.get("purchase_date")
    if purchase_date:
        purchase.purchase_date = date.fromisoformat(purchase_date)

    new_status = request.form.get("status", purchase.status)
    if new_status == "paid" and purchase.status == "due":
        _create_expense_for_purchase(purchase)
        purchase.status = "paid"
    elif new_status == "due" and purchase.status == "paid":
        if purchase.expense_id:
            expense = Expense.query.get(purchase.expense_id)
            if expense:
                db.session.delete(expense)
            purchase.expense_id = None
        purchase.status = "due"
    elif purchase.status == "paid" and purchase.expense_id:
        expense = Expense.query.get(purchase.expense_id)
        if expense:
            expense.amount = purchase.total_amount
            expense.entry_date = purchase.purchase_date
            expense.purpose = f"Purchase - {purchase.supplier.name if purchase.supplier else 'Supplier'}"

    db.session.commit()
    flash("Purchase update kora hoyeche.", "success")
    return redirect(url_for("purchase_page"))


@app.route("/purchases/<int:purchase_id>/delete", methods=["POST"])
@login_required
def delete_purchase(purchase_id):
    purchase = Purchase.query.get_or_404(purchase_id)

    for line in purchase.lines:
        item = Item.query.get(line.item_id)
        if item:
            item.quantity -= line.quantity
        if line.stock_log_id:
            stock_log = StockLog.query.get(line.stock_log_id)
            if stock_log:
                db.session.delete(stock_log)

    if purchase.expense_id:
        expense = Expense.query.get(purchase.expense_id)
        if expense:
            db.session.delete(expense)

    db.session.delete(purchase)
    db.session.commit()
    flash("Purchase entry delete kora hoyeche.", "info")
    return redirect(url_for("purchase_page"))


def _daywise_summary(date_from, date_to):
    income_rows = (
        db.session.query(Income.entry_date, func.sum(Income.amount))
        .filter(Income.entry_date >= date_from, Income.entry_date <= date_to)
        .group_by(Income.entry_date)
        .all()
    )
    expense_rows = (
        db.session.query(Expense.entry_date, func.sum(Expense.amount))
        .filter(Expense.entry_date >= date_from, Expense.entry_date <= date_to)
        .group_by(Expense.entry_date)
        .all()
    )
    income_map = {d: float(amt) for d, amt in income_rows}
    expense_map = {d: float(amt) for d, amt in expense_rows}
    all_dates = sorted(set(income_map) | set(expense_map))

    rows = []
    total_income = 0.0
    total_expense = 0.0
    for d in all_dates:
        inc = income_map.get(d, 0.0)
        exp = expense_map.get(d, 0.0)
        total_income += inc
        total_expense += exp
        rows.append({"date": d, "income": round(inc, 2), "expense": round(exp, 2), "net": round(inc - exp, 2)})

    return rows, round(total_income, 2), round(total_expense, 2), round(total_income - total_expense, 2)


@app.route("/accounts/export")
@login_required
def export_accounts_report():
    today = date.today()
    date_from = date.fromisoformat(request.args.get("date_from") or today.replace(day=1).isoformat())
    date_to = date.fromisoformat(request.args.get("date_to") or today.isoformat())

    daywise_rows, total_income, total_expense, total_net = _daywise_summary(date_from, date_to)

    income_entries = Income.query.filter(
        Income.entry_date >= date_from, Income.entry_date <= date_to
    ).order_by(Income.entry_date, Income.id).all()
    expense_entries = Expense.query.filter(
        Expense.entry_date >= date_from, Expense.entry_date <= date_to
    ).order_by(Expense.entry_date, Expense.id).all()

    transactions = [
        {
            "date": entry.entry_date, "type": "Income", "purpose": entry.purpose,
            "category": entry.category or "", "note": entry.note or "", "amount": entry.amount,
        }
        for entry in income_entries
    ] + [
        {
            "date": entry.entry_date, "type": "Expense", "purpose": entry.purpose,
            "category": entry.category or "", "note": entry.note or "", "amount": entry.amount,
        }
        for entry in expense_entries
    ]
    transactions.sort(key=lambda t: t["date"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    BLUE = "2A78D6"
    RED = "E34948"
    GRAY = "6C757D"
    LIGHT_BLUE = "EAF2FD"
    LIGHT_RED = "FDECEA"
    GOLD = "FFF3CD"
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
    BOLD = Font(bold=True)
    CURRENCY_FMT = '"৳"#,##0.00'
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def section_header(row, text, span=6, fill=GRAY):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = WHITE_FONT
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    def header_row(row, labels):
        for col, label in enumerate(labels, start=1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = BOLD
            cell.fill = PatternFill("solid", fgColor="EEF0F3")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    title_cell = ws.cell(row=r, column=1, value="Income & Expense Report")
    title_cell.font = TITLE_FONT
    title_cell.fill = PatternFill("solid", fgColor=BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 28
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    period_cell = ws.cell(row=r, column=1, value=f"Period: {date_from.isoformat()} to {date_to.isoformat()}")
    period_cell.font = Font(italic=True, color="6C757D")
    r += 2

    section_header(r, "Transaction Details", span=6, fill=GRAY)
    r += 1
    header_row(r, ["Date", "Type", "Purpose", "Category", "Note", "Amount"])
    r += 1
    for t in transactions:
        fill = LIGHT_BLUE if t["type"] == "Income" else LIGHT_RED
        values = [t["date"].isoformat(), t["type"], t["purpose"], t["category"], t["note"], t["amount"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = BORDER
            if col == 2:
                cell.font = Font(color=BLUE if t["type"] == "Income" else RED, bold=True)
            if col == 6:
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right")
        r += 1
    if not transactions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="Ei date range e kono transaction nei.").alignment = Alignment(horizontal="center")
        r += 1
    r += 1

    section_header(r, "Day-wise Totals", span=4, fill=GRAY)
    r += 1
    header_row(r, ["Date", "Income", "Expense", "Net"])
    r += 1
    ZEBRA = "F2F4F7"
    for i, row in enumerate(daywise_rows):
        zebra_fill = PatternFill("solid", fgColor=ZEBRA) if i % 2 == 1 else None
        values = [row["date"].isoformat(), row["income"], row["expense"], row["net"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            if zebra_fill:
                cell.fill = zebra_fill
            if col >= 2:
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right")
            if col == 4:
                cell.font = Font(color="0CA30C" if row["net"] >= 0 else RED, bold=True)
        r += 1
    if not daywise_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=1, value="Ei date range e kono data nei.").alignment = Alignment(horizontal="center")
        r += 1

    total_row = ["All Total", total_income, total_expense, total_net]
    for col, val in enumerate(total_row, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = BOLD
        cell.fill = PatternFill("solid", fgColor=GOLD)
        cell.border = BORDER
        if col >= 2:
            cell.number_format = CURRENCY_FMT
            cell.alignment = Alignment(horizontal="right")
        if col == 4:
            cell.font = Font(bold=True, color="0CA30C" if total_net >= 0 else RED)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A6"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=income_expense_report.xlsx"},
    )


@app.route("/accounts")
@login_required
def accounts_dashboard():
    total_income = db.session.query(func.coalesce(func.sum(Income.amount), 0)).scalar()
    total_expense = db.session.query(func.coalesce(func.sum(Expense.amount), 0)).scalar()
    balance = total_income - total_expense

    today = date.today()
    prev_year, prev_month = (today.year - 1, 12) if today.month == 1 else (today.year, today.month - 1)
    this_month_income = _month_sum(Income, today.year, today.month)
    last_month_income = _month_sum(Income, prev_year, prev_month)
    this_month_expense = _month_sum(Expense, today.year, today.month)
    last_month_expense = _month_sum(Expense, prev_year, prev_month)
    income_trend = _trend(this_month_income, last_month_income)
    expense_trend = _trend(this_month_expense, last_month_expense)

    recent_income = Income.query.order_by(Income.entry_date.desc(), Income.id.desc()).limit(8).all()
    recent_expense = Expense.query.order_by(Expense.entry_date.desc(), Expense.id.desc()).limit(8).all()

    income_by_category = (
        db.session.query(func.coalesce(Income.category, "Uncategorized"), func.sum(Income.amount))
        .group_by(func.coalesce(Income.category, "Uncategorized"))
        .order_by(func.sum(Income.amount).desc())
        .all()
    )
    expense_by_category = (
        db.session.query(func.coalesce(Expense.category, "Uncategorized"), func.sum(Expense.amount))
        .group_by(func.coalesce(Expense.category, "Uncategorized"))
        .order_by(func.sum(Expense.amount).desc())
        .all()
    )

    # Combined monthly summary (last 6 months) so income vs expense is clear at a glance.
    year, month = today.year, today.month
    months = []
    for _ in range(6):
        months.append((year, month))
        year, month = (year - 1, 12) if month == 1 else (year, month - 1)
    months.reverse()

    monthly_summary = []
    max_amount = 0
    for y, m in months:
        m_income = _month_sum(Income, y, m)
        m_expense = _month_sum(Expense, y, m)
        max_amount = max(max_amount, m_income, m_expense)
        monthly_summary.append({
            "label": f"{month_abbr[m]} {y}",
            "income": round(m_income, 2),
            "expense": round(m_expense, 2),
            "net": round(m_income - m_expense, 2),
        })
    for row in monthly_summary:
        row["income_pct"] = round((row["income"] / max_amount) * 100, 1) if max_amount else 0
        row["expense_pct"] = round((row["expense"] / max_amount) * 100, 1) if max_amount else 0

    report_date_from = request.args.get("date_from") or today.replace(day=1).isoformat()
    report_date_to = request.args.get("date_to") or today.isoformat()
    report_date_from_obj = date.fromisoformat(report_date_from)
    report_date_to_obj = date.fromisoformat(report_date_to)
    daywise_rows, period_income, period_expense, period_net = _daywise_summary(
        report_date_from_obj, report_date_to_obj
    )
    report_income_entries = Income.query.filter(
        Income.entry_date >= report_date_from_obj, Income.entry_date <= report_date_to_obj
    ).order_by(Income.entry_date, Income.id).all()
    report_expense_entries = Expense.query.filter(
        Expense.entry_date >= report_date_from_obj, Expense.entry_date <= report_date_to_obj
    ).order_by(Expense.entry_date, Expense.id).all()

    return render_template(
        "accounts_dashboard.html",
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        income_trend=income_trend,
        expense_trend=expense_trend,
        recent_income=recent_income,
        recent_expense=recent_expense,
        income_by_category=income_by_category,
        expense_by_category=expense_by_category,
        monthly_summary=monthly_summary,
        report_date_from=report_date_from,
        report_date_to=report_date_to,
        daywise_rows=daywise_rows,
        period_income=period_income,
        period_expense=period_expense,
        period_net=period_net,
        report_income_entries=report_income_entries,
        report_expense_entries=report_expense_entries,
    )


# ---------- Income ----------

def _filtered_entries(model):
    query = model.query
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    if q:
        query = query.filter(
            db.or_(model.purpose.ilike(f"%{q}%"), model.note.ilike(f"%{q}%"))
        )
    if category:
        query = query.filter(model.category == category)
    if date_from:
        query = query.filter(model.entry_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(model.entry_date <= date.fromisoformat(date_to))

    return query.order_by(model.entry_date.desc(), model.id.desc()).all()


def _distinct_categories(model):
    return [
        row[0] for row in
        db.session.query(model.category).filter(model.category.isnot(None)).distinct().order_by(model.category)
    ]


@app.route("/income")
@login_required
def income():
    entries = _filtered_entries(Income)
    total = sum(entry.amount for entry in entries)
    return render_template(
        "income.html",
        entries=entries,
        total=total,
        categories=_distinct_categories(Income),
        today=date.today().isoformat(),
        filters=request.args,
    )


@app.route("/income/export")
@login_required
def export_income():
    entries = _filtered_entries(Income)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Purpose", "Category", "Amount", "Note"])
    for entry in entries:
        writer.writerow([entry.entry_date, entry.purpose, entry.category or "", entry.amount, entry.note or ""])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=income.csv"},
    )


@app.route("/income/add", methods=["POST"])
@login_required
def add_income():
    purpose = request.form.get("purpose", "").strip()
    category = request.form.get("category", "").strip()
    amount = request.form.get("amount", "0")
    entry_date = request.form.get("entry_date") or date.today().isoformat()
    note = request.form.get("note", "").strip()

    if not purpose or not amount:
        flash("Purpose ar amount dite hobe.", "danger")
        return redirect(url_for("income"))

    entry = Income(
        purpose=purpose,
        category=category or None,
        amount=float(amount),
        entry_date=date.fromisoformat(entry_date),
        note=note,
    )
    db.session.add(entry)
    db.session.commit()
    flash("Income entry add kora hoyeche.", "success")
    return redirect(url_for("income"))


@app.route("/income/<int:entry_id>/delete", methods=["POST"])
@login_required
def delete_income(entry_id):
    entry = Income.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash("Income entry delete kora hoyeche.", "info")
    return redirect(url_for("income"))


# ---------- Expense ----------

@app.route("/expense")
@login_required
def expense():
    entries = _filtered_entries(Expense)
    total = sum(entry.amount for entry in entries)
    return render_template(
        "expense.html",
        entries=entries,
        total=total,
        categories=_distinct_categories(Expense),
        today=date.today().isoformat(),
        filters=request.args,
    )


@app.route("/expense/export")
@login_required
def export_expense():
    entries = _filtered_entries(Expense)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Purpose", "Category", "Amount", "Note"])
    for entry in entries:
        writer.writerow([entry.entry_date, entry.purpose, entry.category or "", entry.amount, entry.note or ""])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=expense.csv"},
    )


@app.route("/expense/add", methods=["POST"])
@login_required
def add_expense():
    purpose = request.form.get("purpose", "").strip()
    category = request.form.get("category", "").strip()
    amount = request.form.get("amount", "0")
    entry_date = request.form.get("entry_date") or date.today().isoformat()
    note = request.form.get("note", "").strip()

    if not purpose or not amount:
        flash("Purpose ar amount dite hobe.", "danger")
        return redirect(url_for("expense"))

    entry = Expense(
        purpose=purpose,
        category=category or None,
        amount=float(amount),
        entry_date=date.fromisoformat(entry_date),
        note=note,
    )
    db.session.add(entry)
    db.session.commit()
    flash("Expense entry add kora hoyeche.", "success")
    return redirect(url_for("expense"))


@app.route("/expense/<int:entry_id>/delete", methods=["POST"])
@login_required
def delete_expense(entry_id):
    entry = Expense.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash("Expense entry delete kora hoyeche.", "info")
    return redirect(url_for("expense"))


def _month_start(year, month):
    return date(year, month, 1)


def _next_month(year, month):
    return (year + 1, 1) if month == 12 else (year, month + 1)


if __name__ == "__main__":
    app.run(debug=True)
